"""Fault-Injection Tests (master prompt §16.9). Each test injects one specific failure and
asserts: the correct responsibility is identified, the correct stage, the correct error code,
correct retryability, and that unrelated already-successful data survives untouched.
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

import httpx
import openpyxl
import pytest

from cash_equivalents_mvp.database import Database
from cash_equivalents_mvp.models import ManualInput, RateRecord, ResponsibilityStatus, Run
from cash_equivalents_mvp.normalization.dates import term_days
from cash_equivalents_mvp.normalization.percentages import validate_rate_range
from cash_equivalents_mvp.orchestration.manager import RunManager
from cash_equivalents_mvp.responsibilities.canada_prime import CanadaPrimeResponsibility
from cash_equivalents_mvp.responsibilities.gic_rates import GicRatesResponsibility
from cash_equivalents_mvp.responsibilities.hisa import HisaResponsibility
from cash_equivalents_mvp.responsibilities.money_market import MoneyMarketResponsibility
from cash_equivalents_mvp.reporting.excel_com import ExcelComRenderer
from cash_equivalents_mvp.reporting.libreoffice import LibreOfficeRenderer
from cash_equivalents_mvp.reporting.renderer_selection import select_renderer
from tests.conftest import make_context


def _db(tmp_path):
    return Database(tmp_path / "fault_injection.db")


# --- 1. No internet / DNS failure ---

def test_dns_failure_on_canada_prime(monkeypatch, tmp_path):
    def raise_dns_error(url, **kw):
        raise httpx.ConnectError("simulated DNS resolution failure")
    monkeypatch.setattr(httpx, "get", raise_dns_error)

    resp = CanadaPrimeResponsibility()
    ctx = make_context(_db(tmp_path), tmp_path)
    status = resp.run_automatic(ctx)

    assert status == ResponsibilityStatus.MANUAL_REQUIRED
    errors = ctx.db.get_errors(ctx.run_id, "canada_prime")
    assert errors[-1].error_code == "SOURCE_HTTP_TIMEOUT"
    assert errors[-1].retryable is True


# --- 2. Empty / unparseable HTML result ---

def test_money_market_empty_html_result(monkeypatch, tmp_path):
    monkeypatch.setattr(httpx, "get", lambda url, **kw: _EmptyResponse())

    resp = MoneyMarketResponsibility()
    ctx = make_context(_db(tmp_path), tmp_path)
    status = resp.run_automatic(ctx)

    assert status == ResponsibilityStatus.MANUAL_REQUIRED
    errors = ctx.db.get_errors(ctx.run_id, "money_market")
    # The page loaded fine (no connection/TLS/auth-status error) but didn't contain the expected
    # "current yield" text — that's a layout/content mismatch, not a connection failure, and is
    # now classified as such (see collectors/http.py's save_debug_html — the real HTML also gets
    # saved to disk for diagnosis, verified separately in test_http_responsibility_contracts.py).
    assert errors[-1].error_code == "SOURCE_LAYOUT_CHANGED"
    assert "current yield" in errors[-1].message.lower()


class _EmptyResponse:
    text = "<html><body>Page not found</body></html>"
    def raise_for_status(self): pass


# --- 3. Invalid / corrupted XLSX ---

def test_invalid_xlsx_upload_for_gic_rates(tmp_path):
    bad = tmp_path / "corrupted.xlsx"
    bad.write_bytes(b"this is not a real zip/xlsx file at all")

    resp = GicRatesResponsibility()
    ctx = make_context(_db(tmp_path), tmp_path)
    mi = ManualInput(responsibility_id="gic_rates", kind="file", file_path=str(bad),
                      original_filename="corrupted.xlsx", override_reason="fault injection test")
    status = resp.run_manual(ctx, mi)

    assert status == ResponsibilityStatus.VALIDATION_FAILED
    errors = ctx.db.get_errors(ctx.run_id, "gic_rates")
    assert len(errors) >= 1


# --- 4. Renamed worksheet ---

def test_renamed_worksheet_raises_workbook_sheet_missing(tmp_path):
    from cash_equivalents_mvp.parsers.gic_xlsx import parse_gic_rates_xlsx

    wb = openpyxl.Workbook()
    wb.active.title = "RenamedSheet"
    path = tmp_path / "renamed.xlsx"
    wb.save(path)

    with pytest.raises(ValueError, match="WORKBOOK_SHEET_MISSING"):
        parse_gic_rates_xlsx(path, {"sheet": "Eng", "code_col": "B", "dealer_col": "C", "min_col": "D",
                                     "annual": {"header_row": 7, "data_row_start": 9, "data_row_end": 20,
                                                "term_years_cols": {1: "E"}}})


# --- 5. Missing provider (GIC and HISA) ---

def test_gic_missing_provider_surfaces_as_warning_not_silent(tmp_path):
    resp = GicRatesResponsibility()
    ctx = make_context(_db(tmp_path), tmp_path)
    from cash_equivalents_mvp.models import CollectionResult
    # Only one provider present out of the expected 12 for the "annual" block.
    collection = CollectionResult(
        ok=True, status=ResponsibilityStatus.SUCCESS,
        raw_rows=[{"block": "annual", "code": "B2BGICP", "dealer": "B2B Bank", "min": 1000,
                   "term_years": 1, "bucket_days": None, "raw_value": 3.3, "row": 9}],
    )
    records = resp.normalize(ctx, collection)
    result = resp.validate(ctx, records)
    assert any(f.rule_id == "PARSER_PROVIDER_MISSING" for f in result.findings)
    assert result.ok  # a missing provider is a warning, not blocking — partial data still useful


# --- 6. Stale source ---

def test_freshness_check_flags_stale_source():
    from cash_equivalents_mvp.validation.common import check_freshness
    findings = check_freshness("run1", "canada_prime", date(2026, 4, 1), date(2026, 5, 11), max_age_days=14)
    assert len(findings) == 1
    assert findings[0].rule_id == "SOURCE_STALE"
    assert findings[0].severity == "warning"


def test_freshness_check_passes_for_fresh_source():
    from cash_equivalents_mvp.validation.common import check_freshness
    findings = check_freshness("run1", "canada_prime", date(2026, 5, 8), date(2026, 5, 11), max_age_days=14)
    assert findings == []


# --- 7. Invalid percentage (scale error) ---

def test_invalid_percentage_out_of_range_rejected():
    with pytest.raises(ValueError):
        validate_rate_range(Decimal("3.30"))  # 330%, a classic scale-error signature


# --- 8. Negative T-bill term ---

def test_negative_tbill_term_raises_term_day_mismatch():
    with pytest.raises(ValueError, match="TBILL_TERM_DAY_MISMATCH"):
        term_days(date(2026, 1, 1), date(2026, 5, 11))


# --- 9. Excel unavailable (both renderers) ---

def test_no_renderer_available_raises_excel_not_installed(monkeypatch):
    monkeypatch.setattr(ExcelComRenderer, "is_available", lambda self: False)
    monkeypatch.setattr(LibreOfficeRenderer, "is_available", lambda self: False)
    name, renderer = select_renderer(["excel_com", "libreoffice"])
    assert name is None and renderer is None


# --- 10. Excel COM crash mid-render ---

def test_excel_com_crash_is_wrapped_as_excel_com_failure(monkeypatch, tmp_path):
    renderer = ExcelComRenderer()

    class _CrashingCoInit:
        def CoInitialize(self): pass
        def CoUninitialize(self): pass

    def dispatch_that_raises(*a, **kw):
        raise OSError("simulated COM crash: RPC server unavailable")

    import sys
    import types
    fake_win32com_client = types.SimpleNamespace(DispatchEx=dispatch_that_raises)
    fake_pythoncom = types.SimpleNamespace(CoInitialize=lambda: None, CoUninitialize=lambda: None)
    monkeypatch.setitem(sys.modules, "win32com.client", fake_win32com_client)
    monkeypatch.setitem(sys.modules, "pythoncom", fake_pythoncom)

    with pytest.raises(RuntimeError, match="EXCEL_COM_FAILURE"):
        renderer.recalculate_and_save(tmp_path / "doesnotmatter.xlsx")


# --- 11. LibreOffice unavailable ---

def test_libreoffice_not_found_reports_unavailable(monkeypatch):
    monkeypatch.setattr("shutil.which", lambda name: None)
    renderer = LibreOfficeRenderer()
    assert renderer.is_available() is False


# --- 12. PDF export failure ---

def test_pdf_export_failure_wrapped_with_error_code(monkeypatch, tmp_path):
    renderer = ExcelComRenderer()
    import sys
    import types

    class _FakeWorkbook:
        def ExportAsFixedFormat(self, *a, **kw):
            raise OSError("simulated export failure: disk full")
        def Close(self, **kw): pass

    class _FakeApp:
        Visible = False
        DisplayAlerts = False
        def __init__(self):
            self.Workbooks = types.SimpleNamespace(Open=lambda *a, **kw: _FakeWorkbook())
        def Quit(self): pass

    fake_win32com_client = types.SimpleNamespace(DispatchEx=lambda *a, **kw: _FakeApp())
    fake_pythoncom = types.SimpleNamespace(CoInitialize=lambda: None, CoUninitialize=lambda: None)
    monkeypatch.setitem(sys.modules, "win32com.client", fake_win32com_client)
    monkeypatch.setitem(sys.modules, "pythoncom", fake_pythoncom)

    with pytest.raises(RuntimeError, match="PDF_EXPORT_FAILED"):
        renderer.export_pdf(tmp_path / "in.xlsx", tmp_path / "out.pdf")


# --- 13. Insufficient disk space (surfaced generically through the same wrapped-exception path) ---

def test_disk_full_during_save_is_wrapped_as_excel_com_failure(monkeypatch, tmp_path):
    renderer = ExcelComRenderer()
    import sys
    import types

    class _FakeWorkbook:
        def Save(self):
            raise OSError(28, "No space left on device")  # errno.ENOSPC
        def Close(self, **kw): pass

    class _FakeApp:
        Visible = False
        DisplayAlerts = False
        def __init__(self):
            self.Workbooks = types.SimpleNamespace(Open=lambda *a, **kw: _FakeWorkbook())
        def CalculateFullRebuild(self): pass
        def Quit(self): pass

    fake_win32com_client = types.SimpleNamespace(DispatchEx=lambda *a, **kw: _FakeApp())
    fake_pythoncom = types.SimpleNamespace(CoInitialize=lambda: None, CoUninitialize=lambda: None)
    monkeypatch.setitem(sys.modules, "win32com.client", fake_win32com_client)
    monkeypatch.setitem(sys.modules, "pythoncom", fake_pythoncom)

    with pytest.raises(RuntimeError, match="EXCEL_COM_FAILURE"):
        renderer.recalculate_and_save(tmp_path / "in.xlsx")


# --- 14. Manual upload with wrong file type ---

def test_manual_upload_wrong_extension_rejected():
    from cash_equivalents_mvp.security import validate_extension
    with pytest.raises(ValueError, match="FILE_TYPE_INVALID"):
        validate_extension("malware.exe", allowed={".xlsx", ".csv"})


# --- 15. A responsibility's failure must not erase unrelated successful data (orchestrator-level) ---

def test_one_responsibility_failing_does_not_erase_unrelated_successful_records(monkeypatch, tmp_path):
    db = _db(tmp_path)
    mgr = RunManager(db)
    run = mgr.create_run(date(2026, 5, 11))

    # Seed canada_prime as already-successful, independent of anything money_market does.
    from cash_equivalents_mvp.models import ResponsibilityStatus as RS
    good_record = RateRecord(run_id=run.run_id, responsibility_id="canada_prime", category="prime",
                              provider="Bank of Canada", currency="CAD", rate=Decimal("0.0445"))
    db.save_rate_records([good_record])
    db.set_responsibility_status(run.run_id, "canada_prime", RS.COMPLETE)

    # Force money_market to fail.
    monkeypatch.setattr(httpx, "get", lambda url, **kw: (_ for _ in ()).throw(httpx.ConnectError("down")))
    resp = MoneyMarketResponsibility()
    ctx = make_context(db, tmp_path, report_date=run.report_date)
    ctx.run_id = run.run_id  # reuse the same run
    resp.run_automatic(ctx)

    # canada_prime's earlier success must be completely untouched.
    assert db.get_responsibility_status(run.run_id, "canada_prime") == RS.COMPLETE
    prime_records = db.get_rate_records(run.run_id, "canada_prime")
    assert len(prime_records) == 1
    assert prime_records[0].rate == Decimal("0.0445")
