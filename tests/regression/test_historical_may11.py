"""Historical Golden Regression (master prompt §17).

Validates against the values baked into the ALREADY-PUBLISHED
`20260511 Cash and Cash Equivalents EN/FR.xlsx` — not a fresh pipeline run against current
(August-dated) sources. This matches the master prompt's own framing ("Use the May 11, 2026
package as a historical fixture only... Do not represent these values as current") and avoids
the trap of asserting that live GIC/T-bill sources will reproduce a frozen historical snapshot,
which they cannot (see docs/source_inventory.md's NBF PDF discussion).

These values are read from the fixture workbook, not hardcoded into production collection logic.
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal

import fitz
import openpyxl
import pytest

from cash_equivalents_mvp.config import settings, source_material_dir

from tests.conftest import requires_source_material

pytestmark = requires_source_material


@pytest.fixture(scope="module")
def en_workbook():
    path = source_material_dir() / settings()["templates"]["en"]
    return openpyxl.load_workbook(path, data_only=True)


@pytest.fixture(scope="module")
def fr_workbook():
    path = source_material_dir() / settings()["templates"]["fr"]
    return openpyxl.load_workbook(path, data_only=True)


def test_report_date_is_may_11_2026(en_workbook):
    assert en_workbook["Cover"]["I29"].value == date(2026, 5, 11).__class__(2026, 5, 11) \
        or en_workbook["Cover"]["I29"].value.date() == date(2026, 5, 11)


def test_canada_prime_4_45_percent(en_workbook):
    assert en_workbook["Cash"]["D31"].value == pytest.approx(0.0445)


def test_us_federal_funds_3_75_percent(en_workbook):
    assert en_workbook["Cash"]["H31"].value == pytest.approx(0.0375)


def test_canadian_money_market_2_00_percent(en_workbook):
    assert "2.00%" in en_workbook["Executive Summary"]["A25"].value


def test_us_money_market_2_82_percent(en_workbook):
    assert "2.82%" in en_workbook["Executive Summary"]["A28"].value


def test_canadian_selected_hisa_2_05_percent(en_workbook):
    assert en_workbook["HISA"]["N13"].value == pytest.approx(0.0205)


def test_us_selected_hisa_3_20_percent(en_workbook):
    assert en_workbook["HISA"]["N17"].value == pytest.approx(0.032)


def test_cashable_gic_bank_of_nova_scotia_1_75_percent(en_workbook):
    ws = en_workbook["Cashable & Term Deposits"]
    assert ws["A57"].value == "Bank of Nova Scotia"
    assert ws["B58"].value == pytest.approx(0.0175)
    # and the underlying row it's drawn from
    assert ws["A21"].value == "Bank of Nova Scotia"
    assert ws["F21"].value == pytest.approx(0.0175)


@pytest.mark.parametrize("cell,expected", [
    ("A56", 0.0205), ("C56", 0.0215), ("E56", 0.0255), ("G56", 0.0260), ("I56", 0.0270), ("K56", 0.0280),
])
def test_term_deposit_rates(en_workbook, cell, expected):
    assert en_workbook["Executive Summary"][cell].value == pytest.approx(expected)


CAD_TBILLS = [
    ("A15", "A19", date(2026, 6, 17), 0.0223),
    ("C15", "C19", date(2026, 7, 15), 0.0223),
    ("E15", "E19", date(2026, 8, 12), 0.0224),
    ("G15", "G19", date(2026, 10, 21), 0.0235),
    ("I15", "I19", date(2027, 4, 21), 0.0259),
]


@pytest.mark.parametrize("date_cell,rate_cell,expected_date,expected_rate", CAD_TBILLS)
def test_canadian_treasury_bills(en_workbook, date_cell, rate_cell, expected_date, expected_rate):
    ws = en_workbook["TBills"]
    assert ws[date_cell].value.date() == expected_date
    assert ws[rate_cell].value == pytest.approx(expected_rate)


US_TBILLS = [
    ("A24", "A28", date(2026, 6, 6), 0.0352),
    ("C24", "C28", date(2026, 7, 14), 0.0351),
    ("E24", "E28", date(2026, 8, 13), 0.0351),
    ("G24", "G28", date(2026, 10, 8), 0.0350),
    ("I24", "I28", date(2027, 4, 15), 0.0352),
]


@pytest.mark.parametrize("date_cell,rate_cell,expected_date,expected_rate", US_TBILLS)
def test_us_treasury_bills(en_workbook, date_cell, rate_cell, expected_date, expected_rate):
    ws = en_workbook["TBills"]
    assert ws[date_cell].value.date() == expected_date
    assert ws[rate_cell].value == pytest.approx(expected_rate)


def test_at_least_one_complete_gic_provider_row(en_workbook):
    ws = en_workbook["GIC 1yr-5yr"]
    assert ws["A18"].value == "Bank of Nova Scotia"
    assert ws["E18"].value == pytest.approx(0.0255)
    assert ws["I18"].value == pytest.approx(0.0305)


def test_at_least_one_complete_hisa_provider_row(en_workbook):
    ws = en_workbook["HISA"]
    assert ws["A30"].value.strip() == "IG Equitable Bank High Interest"
    assert ws["H30"].value == pytest.approx(0.0205)


def test_en_and_fr_report_dates_are_the_same_actual_date(en_workbook, fr_workbook):
    en_date = en_workbook["Cover"]["I29"].value
    fr_date = fr_workbook["Page couverture"]["I29"].value
    assert en_date.date() == fr_date.date() == date(2026, 5, 11)


def test_en_and_fr_gic_1yr5yr_numeric_parity(en_workbook, fr_workbook):
    en_val = en_workbook["GIC 1yr-5yr"]["E18"].value
    fr_val = fr_workbook["CPG 1 an-5 ans"]["E18"].value
    assert en_val == pytest.approx(fr_val)


def test_en_pdf_has_seven_pages():
    path = source_material_dir() / "20260511 Cash and Cash Equivalents EN.pdf"
    doc = fitz.open(path)
    try:
        assert doc.page_count == 7
    finally:
        doc.close()


def test_fr_pdf_has_seven_pages():
    path = source_material_dir() / "20260511 Cash and Cash Equivalents FR.pdf"
    doc = fitz.open(path)
    try:
        assert doc.page_count == 7
    finally:
        doc.close()
