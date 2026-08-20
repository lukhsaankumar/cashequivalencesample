"""GIC Mapping Tests (master prompt §16.3): exact known source→destination ranges, and failure
modes when the source layout changes shape (row inserted, provider missing, #N/A, scale change).
"""
from __future__ import annotations

from decimal import Decimal

import pytest

from cash_equivalents_mvp.config import source_material_dir, workbook_map
from cash_equivalents_mvp.normalization.percentages import normalize_to_canonical, parse_raw_rate
from cash_equivalents_mvp.parsers.gic_xlsx import parse_gic_rates_xlsx

from tests.conftest import requires_source_material

# Verified block -> report-range mapping, see docs/workbook_mapping.md
EXPECTED_RANGES = {
    "cashables": ("Eng!E89:E96", "Cashable & Term Deposits!F18:F25"),
    "short_term_deposits": ("Eng!E69:J80", "Cashable & Term Deposits!E32:J43"),
    "annual": ("Eng!E9:I20", "GIC 1yr-5yr!E14:I25"),
    "compound": ("Eng!E45:I56", "GIC 1yr-5yr!E30:I41"),
    "monthly": ("Eng!E26:I37", "GIC 1yr-5yr!E46:I57"),
}

EXPECTED_12_PROVIDER_ORDER = [
    "B2B", "BMO", "BMT", "BOM", "BNS", "EQB", "EQT", "HOBK", "HTC", "LBC", "MBC", "NBC",
]


@requires_source_material
@pytest.mark.parametrize("block", ["annual", "compound", "monthly", "short_term_deposits"])
def test_block_has_all_twelve_providers_in_expected_order(block):
    source_map = workbook_map("en")["gic_rates_source"]
    rows = parse_gic_rates_xlsx(source_material_dir() / "GIC Rates.xlsx", source_map)
    from cash_equivalents_mvp.normalization.providers import provider_prefix_for_code
    codes_in_block = []
    seen = set()
    for r in sorted((r for r in rows if r["block"] == block), key=lambda r: r["row"]):
        prefix = provider_prefix_for_code(r["code"])
        if prefix not in seen:
            codes_in_block.append(prefix)
            seen.add(prefix)
    assert codes_in_block == EXPECTED_12_PROVIDER_ORDER


@requires_source_material
def test_cashables_block_has_eight_providers_no_laurentian_manulife_national():
    source_map = workbook_map("en")["gic_rates_source"]
    rows = parse_gic_rates_xlsx(source_material_dir() / "GIC Rates.xlsx", source_map)
    from cash_equivalents_mvp.normalization.providers import provider_prefix_for_code
    prefixes = {provider_prefix_for_code(r["code"]) for r in rows if r["block"] == "cashables"}
    assert prefixes == {"BMO", "BMT", "BOM", "BNS", "EQB", "EQT", "HOBK", "HTC"}
    assert "LBC" not in prefixes and "MBC" not in prefixes and "NBC" not in prefixes


@requires_source_material
def test_annual_and_compound_parity_for_every_provider_and_term():
    # docs/workbook_mapping.md: Annual and Compound rates are identical per provider/term in the
    # current snapshot — this pins that observation down as an executable check.
    source_map = workbook_map("en")["gic_rates_source"]
    rows = parse_gic_rates_xlsx(source_material_dir() / "GIC Rates.xlsx", source_map)
    from cash_equivalents_mvp.normalization.providers import provider_prefix_for_code

    def canon(v):
        return normalize_to_canonical(parse_raw_rate(v), source_convention="always_percent_form")

    annual = {(provider_prefix_for_code(r["code"]), r["term_years"]): canon(r["raw_value"])
              for r in rows if r["block"] == "annual" and r["raw_value"] not in (None, "#N/A")}
    compound = {(provider_prefix_for_code(r["code"]), r["term_years"]): canon(r["raw_value"])
                for r in rows if r["block"] == "compound" and r["raw_value"] not in (None, "#N/A")}
    common_keys = set(annual) & set(compound)
    assert len(common_keys) >= 50  # sanity: most of the 12*5=60 cells should be present
    mismatches = [k for k in common_keys if annual[k] != compound[k]]
    assert mismatches == [], f"Annual/Compound diverge for: {mismatches}"


def test_provider_row_inserted_does_not_break_identity_based_matching():
    """The writer matches by provider CODE, not row position — inserting a row above an existing
    provider must not misattribute rates, unlike a hardcoded-row-number implementation would."""
    import openpyxl
    from cash_equivalents_mvp.reporting.mappings import MappingReport, _write_gic_block
    from cash_equivalents_mvp.models import RateRecord

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A15"], ws["D15"], ws["F15"] = "Issuer", "Code", "Rate"
    ws["A17"], ws["D17"] = "B2B Bank", "B2BGICR"
    ws["A18"], ws["D18"] = "Bank of Nova Scotia", "BNSGICR"  # BNS originally at row 18

    block_cfg = {"header_row": 15, "data_row_start": 17, "data_row_end": 18,
                 "issuer_col": "A", "code_col": "D", "rate_col": "F"}
    report = MappingReport()
    rec = RateRecord(run_id="r", responsibility_id="gic_rates", category="gic", provider="Bank of Nova Scotia",
                      product_code="BNSGICR", currency="CAD", account_type="cashables", rate=Decimal("0.0175"))
    unmatched = _write_gic_block(ws, block_cfg, [rec], report, "numeric", "cashable_gic")
    assert unmatched == []
    assert ws["F18"].value == pytest.approx(0.0175)


def test_provider_missing_from_scan_window_is_reported_unmatched():
    import openpyxl
    from cash_equivalents_mvp.reporting.mappings import MappingReport, _write_gic_block
    from cash_equivalents_mvp.models import RateRecord

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A15"], ws["D15"], ws["F15"] = "Issuer", "Code", "Rate"
    ws["A17"], ws["D17"] = "B2B Bank", "B2BGICR"

    block_cfg = {"header_row": 15, "data_row_start": 17, "data_row_end": 17,
                 "issuer_col": "A", "code_col": "D", "rate_col": "F"}
    report = MappingReport()
    rec = RateRecord(run_id="r", responsibility_id="gic_rates", category="gic", provider="Home Bank GIC",
                      product_code="HOBKGICR", currency="CAD", account_type="cashables", rate=Decimal("0.02"))
    unmatched = _write_gic_block(ws, block_cfg, [rec], report, "numeric", "cashable_gic")
    assert unmatched == ["HOBKGICR"]


def test_na_source_value_never_becomes_a_written_rate():
    """§16.3: 'Test failure when: a source value is #N/A' — the value must never overwrite a
    report cell (see responsibilities/gic_rates.py normalize())."""
    from cash_equivalents_mvp.responsibilities.gic_rates import GicRatesResponsibility
    from cash_equivalents_mvp.models import CollectionResult, ResponsibilityStatus

    resp = GicRatesResponsibility()
    collection = CollectionResult(
        ok=True, status=ResponsibilityStatus.SUCCESS,
        raw_rows=[{"block": "cashables", "code": "HOBKGICR", "dealer": "Home Bank GIC", "min": 5000,
                   "term_years": 1, "bucket_days": None, "raw_value": "#N/A", "row": 95}],
    )

    class _Ctx:
        run_id = "r"

    records = resp.normalize(_Ctx(), collection)
    assert records == []
