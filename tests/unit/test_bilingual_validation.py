import openpyxl

from cash_equivalents_mvp.validation.bilingual import check_bilingual_parity


def _make_minimal_workbook(tmp_path, name, sheets: dict[str, dict[str, object]]):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, cells in sheets.items():
        ws = wb.create_sheet(sheet_name)
        for coord, value in cells.items():
            ws[coord] = value
    path = tmp_path / name
    wb.save(path)
    return path


def _minimal_wmap(cash_sheet, prime_cell, fed_cell, prime_fmt, fed_fmt, es_sheet):
    return {
        "sheets": {"cash": cash_sheet, "executive_summary": es_sheet},
        "prime": {"sheet": "cash", "value_cell": prime_cell, "value_format": prime_fmt,
                  "label_cell": "A1", "label_text": ""},
        "fed_funds": {"sheet": "cash", "value_cell": fed_cell, "value_format": fed_fmt,
                      "label_cell": "A1", "label_text": ""},
        "money_market_text": {"sheet": "executive_summary", "cad_cell": "A1", "us_cell": "A2",
                               "template": "{rate}"},
    }


def test_matching_values_pass_parity_check(tmp_path):
    en_map = _minimal_wmap("Cash", "D31", "H31", "numeric", "numeric", "Summary")
    fr_map = _minimal_wmap("Espèces", "E32", "H32", "french_percent_text", "french_percent_text", "Sommaire")
    en_path = _make_minimal_workbook(tmp_path, "en.xlsx", {
        "Cash": {"D31": 0.0445, "H31": 0.0375}, "Summary": {"A1": "2.00%", "A2": "2.82%"},
    })
    fr_path = _make_minimal_workbook(tmp_path, "fr.xlsx", {
        "Espèces": {"E32": "4,45%", "H32": "3,75%"}, "Sommaire": {"A1": "2%", "A2": "2,82%"},
    })
    result = check_bilingual_parity(en_path, fr_path, en_map, fr_map)
    assert result.ok, result.mismatches


def test_mismatched_prime_rate_fails_parity_check(tmp_path):
    en_map = _minimal_wmap("Cash", "D31", "H31", "numeric", "numeric", "Summary")
    fr_map = _minimal_wmap("Espèces", "E32", "H32", "french_percent_text", "french_percent_text", "Sommaire")
    en_path = _make_minimal_workbook(tmp_path, "en.xlsx", {
        "Cash": {"D31": 0.0445, "H31": 0.0375}, "Summary": {"A1": "2.00%", "A2": "2.82%"},
    })
    fr_path = _make_minimal_workbook(tmp_path, "fr.xlsx", {
        # deliberately stale FR prime rate — simulates the exact bug class this check exists to catch
        "Espèces": {"E32": "4,25%", "H32": "3,75%"}, "Sommaire": {"A1": "2%", "A2": "2,82%"},
    })
    result = check_bilingual_parity(en_path, fr_path, en_map, fr_map)
    assert not result.ok
    assert any("prime" in m for m in result.mismatches)
    assert not any("fed_funds" in m for m in result.mismatches)  # unaffected cell must not false-positive
