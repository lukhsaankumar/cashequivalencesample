"""Parses GIC Rates.xlsx (the 'Eng' sheet) using the verified block layout in
config/workbook_map_en.yaml: gic_rates_source. See docs/source_inventory.md and
docs/workbook_mapping.md for how each block/range was confirmed against the real file.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

PARSER_VERSION = "gic_xlsx-1.0"


def _col_letter_to_index(letter: str) -> int:
    from openpyxl.utils import column_index_from_string
    return column_index_from_string(letter)


def parse_gic_rates_xlsx(path: Path, source_map: dict) -> list[dict[str, Any]]:
    """Returns raw rows (no scaling/normalization applied yet — that's the responsibility's job).

    Each row: {block, code, dealer, min, term_years|None, bucket_days|None, raw_value}
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = source_map["sheet"]
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"WORKBOOK_SHEET_MISSING: sheet {sheet_name!r} not found in {path.name}")
    ws = wb[sheet_name]

    code_col = source_map["code_col"]
    dealer_col = source_map["dealer_col"]
    min_col = source_map["min_col"]

    rows: list[dict[str, Any]] = []

    for block_name, key in [("annual", "annual"), ("monthly", "monthly"), ("compound", "compound")]:
        block = source_map.get(key)
        if not block:
            continue
        for row_idx in range(block["data_row_start"], block["data_row_end"] + 1):
            code = ws[f"{code_col}{row_idx}"].value
            dealer = ws[f"{dealer_col}{row_idx}"].value
            if not code or not str(code).strip():
                continue
            min_val = ws[f"{min_col}{row_idx}"].value
            for term_years, col in block["term_years_cols"].items():
                raw = ws[f"{col}{row_idx}"].value
                rows.append({
                    "block": block_name,
                    "code": str(code).strip(),
                    "dealer": str(dealer).strip() if dealer else None,
                    "min": min_val,
                    "term_years": int(term_years),
                    "bucket_days": None,
                    "raw_value": raw,
                    "row": row_idx,
                })

    std = source_map.get("short_term_deposits")
    if std:
        for row_idx in range(std["data_row_start"], std["data_row_end"] + 1):
            code = ws[f"{code_col}{row_idx}"].value
            dealer = ws[f"{dealer_col}{row_idx}"].value
            if not code or not str(code).strip():
                continue
            min_val = ws[f"{min_col}{row_idx}"].value
            for bucket_days, col in std["buckets"].items():
                raw = ws[f"{col}{row_idx}"].value
                rows.append({
                    "block": "short_term_deposits",
                    "code": str(code).strip(),
                    "dealer": str(dealer).strip() if dealer else None,
                    "min": min_val,
                    "term_years": None,
                    "bucket_days": int(bucket_days),
                    "raw_value": raw,
                    "row": row_idx,
                })

    cashables = source_map.get("cashables")
    if cashables:
        for row_idx in range(cashables["data_row_start"], cashables["data_row_end"] + 1):
            code = ws[f"{code_col}{row_idx}"].value
            dealer = ws[f"{dealer_col}{row_idx}"].value
            if not code or not str(code).strip():
                continue
            min_val = ws[f"{min_col}{row_idx}"].value
            raw = ws[f"{cashables['rate_col']}{row_idx}"].value
            rows.append({
                "block": "cashables",
                "code": str(code).strip(),
                "dealer": str(dealer).strip() if dealer else None,
                "min": min_val,
                "term_years": 1,
                "bucket_days": None,
                "raw_value": raw,
                "row": row_idx,
            })

    if not rows:
        raise ValueError("PARSER_NO_ROWS: no GIC rate rows found — layout may have changed")

    return rows
