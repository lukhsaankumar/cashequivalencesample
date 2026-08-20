"""HISA parsers.

parse_hisa_from_workbook reads the detailed provider table directly out of a report workbook's
HISA sheet (config/workbook_map_en.yaml: hisa) — used as the historical-demo "current roster"
fixture, exactly as gic_rates.py reuses GIC Rates.xlsx from source_material/. This is a legitimate
stand-in for what would otherwise be a scraped/manually-curated roster: in production the same
normalized shape would come from the authenticated home.investorsgroup.com pages or a manual
upload matching parse_hisa_csv's canonical schema.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

PARSER_VERSION = "hisa-1.0"

CANONICAL_CSV_COLUMNS = [
    "provider", "fund_code", "minimum", "maximum", "corporate_eligible", "cdic_eligible",
    "currency", "rate", "source", "effective_date",
]


def parse_hisa_from_workbook(path: Path, hisa_map: dict, sheet_name: str) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"WORKBOOK_SHEET_MISSING: sheet {sheet_name!r} not found in {path.name}")
    ws = wb[sheet_name]

    rows: list[dict[str, Any]] = []

    def _read_block(row_start: int, row_end: int, currency: str) -> None:
        for r in range(row_start, row_end + 1):
            issuer = ws[f"{hisa_map['issuer_col']}{r}"].value
            fund_code = ws[f"{hisa_map['fund_code_col']}{r}"].value
            rate = ws[f"{hisa_map['yield_col']}{r}"].value
            if not issuer or rate is None:
                continue
            rows.append({
                "provider": str(issuer).strip(),
                "fund_code": str(fund_code).strip() if fund_code else None,
                "minimum": ws[f"{hisa_map['minimum_col']}{r}"].value if "minimum_col" in hisa_map else None,
                "maximum": ws[f"{hisa_map['maximum_col']}{r}"].value if "maximum_col" in hisa_map else None,
                "corporate_eligible": ws[f"{hisa_map['corporate_eligible_col']}{r}"].value if "corporate_eligible_col" in hisa_map else None,
                "cdic_eligible": ws[f"{hisa_map['cdic_col']}{r}"].value if "cdic_col" in hisa_map else None,
                "currency": currency,
                "raw_rate": rate,
                "row": r,
            })

    if "cdn_data_row_start" in hisa_map:
        _read_block(hisa_map["cdn_data_row_start"], hisa_map["cdn_data_row_end"], "CAD")
        _read_block(hisa_map["us_data_row_start"], hisa_map["us_data_row_end"], "USD")
    else:
        # FR-style generous scan window with no known CDN/US split — classify by presence of "US$"/"USD" in provider name
        for r in range(hisa_map["scan_row_start"], hisa_map["scan_row_end"] + 1):
            issuer = ws[f"{hisa_map['issuer_col']}{r}"].value
            fund_code = ws[f"{hisa_map.get('fund_code_col', 'C')}{r}"].value
            rate = ws[f"{hisa_map['yield_col']}{r}"].value
            if not issuer or rate is None:
                continue
            currency = "USD" if any(tok in str(issuer).upper() for tok in ("US$", "USD", " US")) else "CAD"
            rows.append({"provider": str(issuer).strip(), "fund_code": str(fund_code).strip() if fund_code else None,
                          "minimum": None, "maximum": None, "corporate_eligible": None, "cdic_eligible": None,
                          "currency": currency, "raw_rate": rate, "row": r})

    if not rows:
        raise ValueError("PARSER_NO_ROWS: no HISA rows found — layout may have changed")
    return rows


def parse_hisa_csv(path: Path) -> list[dict[str, Any]]:
    import csv
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None or not set(CANONICAL_CSV_COLUMNS).issubset({h.strip() for h in reader.fieldnames}):
            raise ValueError(f"PARSER_NO_ROWS: HISA CSV must have columns {CANONICAL_CSV_COLUMNS}")
        rows = []
        for line in reader:
            provider = (line.get("provider") or "").strip()
            if not provider:
                continue
            rows.append({
                "provider": provider,
                "fund_code": (line.get("fund_code") or "").strip() or None,
                "minimum": line.get("minimum") or None,
                "maximum": line.get("maximum") or None,
                "corporate_eligible": line.get("corporate_eligible"),
                "cdic_eligible": line.get("cdic_eligible"),
                "currency": (line.get("currency") or "CAD").strip().upper(),
                "raw_rate": line.get("rate"),
                "row": None,
            })
    if not rows:
        raise ValueError("PARSER_NO_ROWS: HISA CSV contained no data rows")
    return rows
