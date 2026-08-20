"""Generic engine that writes canonical RateRecords into a workbook copy using
config/workbook_map_{en,fr}.yaml. Only ever sets `.cell.value` — never touches formulas,
formatting, merged cells, print areas, or conditional formatting, all of which openpyxl leaves
alone unless explicitly written to. Excel recalculation happens later, via COM
(reporting/excel_com.py) — openpyxl cannot recalculate formulas itself.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from cash_equivalents_mvp.models import RateRecord
from cash_equivalents_mvp.normalization.percentages import french_percent_text, french_percent_text_trim_zero
from cash_equivalents_mvp.normalization.providers import provider_prefix_for_code


class WorkbookMappingError(Exception):
    def __init__(self, error_code: str, message: str):
        self.error_code = error_code
        super().__init__(f"{error_code}: {message}")


@dataclass
class MappingReport:
    writes: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def write(self, sheet: str, cell: str, old, new) -> None:
        self.writes.append({"sheet": sheet, "cell": cell, "old": old, "new": new})


def _fmt(value: Decimal, value_format: str) -> object:
    if value_format == "numeric":
        return float(value)
    if value_format == "french_percent_text":
        return french_percent_text(value)
    if value_format == "french_percent_text_trim_zero":
        return french_percent_text_trim_zero(value)
    raise ValueError(f"Unknown value_format {value_format!r}")


def _sheet(wb: openpyxl.Workbook, wmap: dict, section_key: str, language: str) -> Worksheet:
    alias = wmap[section_key]["sheet"]
    name = wmap["sheets"][alias]
    if name not in wb.sheetnames:
        raise WorkbookMappingError("WORKBOOK_SHEET_MISSING", f"{language}: sheet {name!r} not found")
    return wb[name]


def write_report_date(wb: openpyxl.Workbook, wmap: dict, language: str, report_date, report: MappingReport) -> None:
    ws = _sheet(wb, wmap, "report_date", language)
    cell = wmap["report_date"]["cell"]
    old = ws[cell].value
    ws[cell] = report_date
    report.write(ws.title, cell, old, report_date)


def write_prime(wb: openpyxl.Workbook, wmap: dict, language: str, rate: Decimal, report: MappingReport) -> None:
    cfg = wmap["prime"]
    ws = _sheet(wb, wmap, "prime", language)
    label = str(ws[cfg["label_cell"]].value or "")
    if cfg["label_text"].lower() not in label.lower():
        raise WorkbookMappingError("WORKBOOK_LABEL_MISMATCH",
                                    f"{language} Cash!{cfg['label_cell']} expected label containing "
                                    f"{cfg['label_text']!r}, found {label!r}")
    value_format = cfg.get("value_format", "numeric")
    new_val = _fmt(rate, value_format)
    old = ws[cfg["value_cell"]].value
    ws[cfg["value_cell"]] = new_val
    report.write(ws.title, cfg["value_cell"], old, new_val)


def write_fed_funds(wb: openpyxl.Workbook, wmap: dict, language: str, rate: Decimal, report: MappingReport) -> None:
    cfg = wmap["fed_funds"]
    ws = _sheet(wb, wmap, "fed_funds", language)
    label = str(ws[cfg["label_cell"]].value or "")
    if cfg["label_text"].lower() not in label.lower():
        raise WorkbookMappingError("WORKBOOK_LABEL_MISMATCH",
                                    f"{language} Cash!{cfg['label_cell']} expected label containing "
                                    f"{cfg['label_text']!r}, found {label!r}")
    value_format = cfg.get("value_format", "numeric")
    new_val = _fmt(rate, value_format)
    old = ws[cfg["value_cell"]].value
    ws[cfg["value_cell"]] = new_val
    report.write(ws.title, cfg["value_cell"], old, new_val)


def write_money_market_text(wb: openpyxl.Workbook, wmap: dict, language: str,
                             cad_rate: Decimal, us_rate: Decimal, report: MappingReport) -> None:
    cfg = wmap["money_market_text"]
    ws = _sheet(wb, wmap, "money_market_text", language)
    value_format = cfg.get("value_format", "french_percent_text_trim_zero" if language == "fr" else "en_percent")
    for cell_key, rate in [("cad_cell", cad_rate), ("us_cell", us_rate)]:
        cell = cfg[cell_key]
        if language == "fr":
            rate_str = _fmt(rate, value_format)
        else:
            from cash_equivalents_mvp.normalization.percentages import en_percent_string
            rate_str = en_percent_string(rate)
        new_val = cfg["template"].format(rate=rate_str)
        old = ws[cell].value
        ws[cell] = new_val
        report.write(ws.title, cell, old, new_val)


def write_tbills(wb: openpyxl.Workbook, wmap: dict, language: str, records: list[RateRecord],
                  report: MappingReport) -> None:
    cfg = wmap["tbills"]
    ws = _sheet(wb, wmap, "tbills", language)
    value_format = cfg.get("value_format", "numeric")
    for currency, block_key in [("CAD", "canadian"), ("USD", "us")]:
        block = cfg[block_key]
        # T-bill records always carry a real maturity_date (set by treasury_bills.py's
        # normalize()); date.max is just a type-safe fallback, never actually hit.
        currency_records = sorted([r for r in records if r.currency == currency],
                                   key=lambda r: r.maturity_date or date.max)
        slots = min(len(currency_records), len(block["maturity_cells"]))
        for i in range(slots):
            r = currency_records[i]
            mat_cell = block["maturity_cells"][i]
            rate_cell = block["rate_cells"][i]
            old_m = ws[mat_cell].value
            ws[mat_cell] = r.maturity_date
            report.write(ws.title, mat_cell, old_m, r.maturity_date)
            new_rate = _fmt(r.rate, value_format)
            old_r = ws[rate_cell].value
            ws[rate_cell] = new_rate
            report.write(ws.title, rate_cell, old_r, new_rate)
        if len(currency_records) != len(block["maturity_cells"]):
            report.warnings.append(
                f"{language} TBills {currency}: expected {len(block['maturity_cells'])} maturities, "
                f"got {len(currency_records)}"
            )


def _write_gic_block(ws: Worksheet, block_cfg: dict, records: list[RateRecord], report: MappingReport,
                      value_format: str, sheet_label: str) -> list[str]:
    unmatched: list[str] = []
    code_to_row: dict[str, int] = {}
    for row in range(block_cfg["data_row_start"], block_cfg["data_row_end"] + 1):
        code = ws[f"{block_cfg['code_col']}{row}"].value
        if code:
            code_to_row[str(code).strip().upper()] = row

    for rec in records:
        code = (rec.product_code or "").strip().upper()
        target_row = code_to_row.get(code)
        if target_row is None:
            # fall back to provider-prefix match (handles R/P/G suffix differences between
            # GIC Rates.xlsx codes and the report workbook's own D-column codes)
            prefix = provider_prefix_for_code(code)
            for existing_code, row in code_to_row.items():
                if provider_prefix_for_code(existing_code) == prefix and prefix is not None:
                    target_row = row
                    break
        if target_row is None:
            unmatched.append(code)
            continue
        col = (block_cfg.get("rate_col") or
               block_cfg.get("term_years_cols", {}).get(rec.term_years) or
               block_cfg.get("buckets", {}).get(rec.term_days))
        if col is None:
            continue
        cell = f"{col}{target_row}"
        new_val = _fmt(rec.rate, value_format)
        old = ws[cell].value
        ws[cell] = new_val
        report.write(ws.title, cell, old, new_val)
    return unmatched


def write_cashable_and_term_deposits(wb: openpyxl.Workbook, wmap: dict, language: str,
                                      records: list[RateRecord], report: MappingReport) -> None:
    cashable_cfg = wmap["cashable_gic"]
    term_cfg = wmap["term_deposits"]
    ws = _sheet(wb, wmap, "cashable_gic", language)
    value_format = cashable_cfg.get("value_format", "numeric")

    cashable_records = [r for r in records if r.account_type == "cashables"]
    term_records = [r for r in records if r.account_type == "short_term_deposits"]

    unmatched = _write_gic_block(ws, cashable_cfg, cashable_records, report, value_format, "cashable_gic")
    unmatched += _write_gic_block(ws, term_cfg, term_records, report, value_format, "term_deposits")
    for code in unmatched:
        report.warnings.append(f"{language} Cashable/Term Deposits: no matching row for provider code {code!r}")


def write_gic_1yr_5yr(wb: openpyxl.Workbook, wmap: dict, language: str, records: list[RateRecord],
                       report: MappingReport) -> None:
    ws = _sheet(wb, wmap, "gic_annual", language)
    value_format = wmap["gic_annual"].get("value_format", "numeric")
    unmatched = []
    for block_key, account_type in [("gic_annual", "annual"), ("gic_compound", "compound"), ("gic_monthly", "monthly")]:
        block_cfg = wmap[block_key]
        block_records = [r for r in records if r.account_type == account_type]
        unmatched += _write_gic_block(ws, block_cfg, block_records, report, value_format, block_key)
    for code in unmatched:
        report.warnings.append(f"{language} GIC 1yr-5yr: no matching row for provider code {code!r}")


def write_hisa(wb: openpyxl.Workbook, wmap: dict, language: str, records: list[RateRecord],
                report: MappingReport) -> None:
    cfg = wmap["hisa"]
    ws = _sheet(wb, wmap, "hisa", language)
    value_format = cfg.get("value_format", "numeric")

    fund_code_to_row: dict[str, int] = {}
    row_range: list[int]
    if "discovery" in cfg and cfg["discovery"] == "scan":
        row_range = list(range(cfg["scan_row_start"], cfg["scan_row_end"] + 1))
    else:
        row_range = list(range(cfg["cdn_data_row_start"], cfg["cdn_data_row_end"] + 1)) + \
                    list(range(cfg["us_data_row_start"], cfg["us_data_row_end"] + 1))
    for row in row_range:
        code = ws[f"{cfg['fund_code_col']}{row}"].value
        if code:
            fund_code_to_row[str(code).strip().upper()] = row

    unmatched: list[str] = []
    for rec in records:
        code = (rec.product_code or "").strip().upper()
        matched_row = fund_code_to_row.get(code)
        if matched_row is None:
            unmatched.append(f"{rec.provider}/{rec.product_code}")
            continue
        cell = f"{cfg['yield_col']}{matched_row}"
        new_val = _fmt(rec.rate, value_format)
        old = ws[cell].value
        ws[cell] = new_val
        report.write(ws.title, cell, old, new_val)
    for u in unmatched:
        report.warnings.append(f"{language} HISA: no matching row for {u}")
