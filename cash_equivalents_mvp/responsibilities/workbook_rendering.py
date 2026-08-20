"""Workbook Rendering responsibility — populates the EN and FR working copies from every
collector responsibility's canonical RateRecords, then recalculates through Excel (or LibreOffice)
and saves. This is also where French Output (master prompt §7.9) happens: both languages are
written from the exact same canonical Decimal in the same pass, so bilingual numeric parity is
guaranteed by construction rather than needing a separate reconciliation responsibility — see
ASSUMPTIONS.md for why this wasn't split into two DAG nodes.

Reuses the Responsibility interface loosely: this stage doesn't produce RateRecords (it consumes
every other responsibility's), so normalize() does the actual populate+recalculate work as a side
effect and returns [], while validate() re-opens the saved files to confirm the write landed and
formulas/print areas survived.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from cash_equivalents_mvp.config import settings, workbook_map
from cash_equivalents_mvp.models import (
    CollectionResult,
    ManualInput,
    RateRecord,
    ResponsibilityError,
    ResponsibilityStatus,
    ValidationResult,
)
from cash_equivalents_mvp.reporting import mappings
from cash_equivalents_mvp.reporting.renderer_selection import select_renderer
from cash_equivalents_mvp.responsibilities.base import Responsibility, RunContext
from cash_equivalents_mvp.validation.bilingual import check_bilingual_parity
from cash_equivalents_mvp.validation.common import finding

FORMULA_ERROR_TOKENS = ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A", "#NULL!", "#NUM!")


class WorkbookRenderingResponsibility(Responsibility):
    responsibility_id = "workbook_rendering"
    display_name = "Workbook Rendering (EN + FR)"
    dependencies = ("template", "report_date", "gic_rates", "canada_prime", "us_fed_funds",
                     "money_market", "treasury_bills", "hisa")

    def collect_automatic(self, context: RunContext) -> CollectionResult:
        paths_file = context.run_dir / "template_paths.txt"
        if not paths_file.exists():
            err = ResponsibilityError(run_id=context.run_id, responsibility_id=self.responsibility_id,
                                       stage="collect_automatic", error_code="FILE_MISSING",
                                       message="Working template copies not found — Template responsibility "
                                               "must complete first.")
            return CollectionResult(ok=False, status=ResponsibilityStatus.BLOCKED, error=err)
        en_path, fr_path = (Path(p) for p in paths_file.read_text(encoding="utf-8").splitlines())
        return CollectionResult(ok=True, status=ResponsibilityStatus.SUCCESS,
                                 raw_rows=[{"en": str(en_path), "fr": str(fr_path)}])

    def parse_manual_input(self, context: RunContext, manual_input: ManualInput) -> CollectionResult:
        return self.collect_automatic(context)  # no separate manual data path — this is a "re-render" trigger

    def normalize(self, context: RunContext, collection: CollectionResult) -> list[RateRecord]:
        en_path, fr_path = Path(collection.raw_rows[0]["en"]), Path(collection.raw_rows[0]["fr"])

        records_by_resp = {
            rid: context.db.get_rate_records(context.run_id, rid)
            for rid in ("gic_rates", "canada_prime", "us_fed_funds", "money_market", "treasury_bills", "hisa")
        }
        all_gic = records_by_resp["gic_rates"]
        prime = records_by_resp["canada_prime"][0].rate if records_by_resp["canada_prime"] else None
        fed = records_by_resp["us_fed_funds"][0].rate if records_by_resp["us_fed_funds"] else None
        mm = {r.currency: r.rate for r in records_by_resp["money_market"]}
        tbills = records_by_resp["treasury_bills"]
        hisa = records_by_resp["hisa"]

        for language, path in [("en", en_path), ("fr", fr_path)]:
            wmap = workbook_map(language)
            wb = openpyxl.load_workbook(path)
            report = mappings.MappingReport()

            mappings.write_report_date(wb, wmap, language, context.report_date, report)
            if prime is not None:
                mappings.write_prime(wb, wmap, language, prime, report)
            if fed is not None:
                mappings.write_fed_funds(wb, wmap, language, fed, report)
            if "CAD" in mm and "USD" in mm:
                mappings.write_money_market_text(wb, wmap, language, mm["CAD"], mm["USD"], report)
            mappings.write_tbills(wb, wmap, language, tbills, report)
            mappings.write_cashable_and_term_deposits(wb, wmap, language, all_gic, report)
            mappings.write_gic_1yr_5yr(wb, wmap, language, all_gic, report)
            mappings.write_hisa(wb, wmap, language, hisa, report)

            wb.save(path)

            (context.run_dir / f"mapping_report_{language}.json").write_text(
                __import__("json").dumps({"writes": len(report.writes), "warnings": report.warnings},
                                          default=str, indent=2),
                encoding="utf-8",
            )

        renderer_name, renderer = select_renderer(settings()["renderer"]["preference"])
        if renderer is None:
            raise RuntimeError("EXCEL_NOT_INSTALLED: neither Microsoft Excel nor LibreOffice is available")

        for path in (en_path, fr_path):
            renderer.recalculate_and_save(path)

        (context.run_dir / "renderer.txt").write_text(renderer_name or "unknown", encoding="utf-8")
        return []

    def validate(self, context: RunContext, records: list[RateRecord]) -> ValidationResult:
        findings = []
        rid = self.responsibility_id
        paths_file = context.run_dir / "template_paths.txt"
        if not paths_file.exists():
            findings.append(finding(context.run_id, rid, "FILE_MISSING", "blocking", "No rendered workbooks found"))
            return ValidationResult(ok=False, findings=findings)

        # Scoped to print areas, not the whole workbook: hidden calculator/scratch sheets
        # (Bankers Acceptance, IGSI F.I. Calculator, Par Value Finder) and off-print-area scratch
        # columns carry pre-existing formula errors in the *original, untouched* template —
        # verified via scripts/check_pristine_errors.py — and are never part of the printed
        # deliverable, so flagging them here would block every run on a defect the pipeline
        # didn't cause and can't fix. The PDF Export responsibility separately scans the actual
        # rendered PDF text (i.e. exactly what print areas produce) per master prompt §7.11.
        en_path, fr_path = (Path(p) for p in paths_file.read_text(encoding="utf-8").splitlines())
        for language, path in [("en", en_path), ("fr", fr_path)]:
            wb = openpyxl.load_workbook(path, data_only=True)
            for ws in wb.worksheets:
                if ws.sheet_state != "visible":
                    continue  # hidden calculator/scratch sheets are never exported to the PDF
                bounds = _print_area_bounds(ws)
                if bounds is None:
                    continue
                min_col, min_row, max_col, max_row = bounds
                for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                    for cell in row:
                        if isinstance(cell.value, str) and any(tok in cell.value for tok in FORMULA_ERROR_TOKENS):
                            findings.append(finding(context.run_id, rid, "WORKBOOK_FORMULA_OVERWRITTEN", "blocking",
                                                     f"{language} {ws.title}!{cell.coordinate} (print area) "
                                                     f"contains a formula error: {cell.value!r}"))

        parity = check_bilingual_parity(en_path, fr_path, wmap_en=workbook_map("en"), wmap_fr=workbook_map("fr"))
        for mismatch in parity.mismatches:
            findings.append(finding(context.run_id, rid, "BILINGUAL_PARITY_FAILED", "blocking", mismatch))

        return ValidationResult(ok=not any(f.severity == "blocking" for f in findings), findings=findings)


def _print_area_bounds(ws) -> tuple[int, int, int, int] | None:
    if not ws.print_area:
        return None
    from openpyxl.utils import range_boundaries
    try:
        first_range = ws.print_area[0] if isinstance(ws.print_area, list) else ws.print_area
        min_col, min_row, max_col, max_row = range_boundaries(first_range.split("!")[-1])
        return min_col, min_row, max_col, max_row
    except Exception:
        return None
