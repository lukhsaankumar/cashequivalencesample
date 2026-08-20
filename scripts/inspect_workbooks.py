"""One-off inspection script: dumps workbook structure to help build docs/workbook_mapping.md.
Not part of the application; used during bootstrap only.
"""
import sys
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "source_material"


def inspect(path: Path, max_rows=100, max_cols=15):
    print("=" * 100)
    print("FILE:", path.name)
    wb = openpyxl.load_workbook(path, data_only=False)
    print("Sheets:", wb.sheetnames)
    for name in wb.sheetnames:
        ws = wb[name]
        print("-" * 80)
        print(f"SHEET: {name!r} visible={ws.sheet_state} dims={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column}")
        if ws.print_area:
            print("  print_area:", ws.print_area)
        if ws.merged_cells.ranges:
            print("  merged_cells:", list(ws.merged_cells.ranges)[:40])
        if ws.conditional_formatting:
            cf_ranges = [str(r) for r in ws.conditional_formatting]
            if cf_ranges:
                print("  conditional_formatting ranges:", cf_ranges[:20])
        try:
            defined = [dn for dn in wb.defined_names if True]
        except Exception:
            defined = []
    if wb.defined_names:
        print("Defined names:", list(wb.defined_names.keys()) if hasattr(wb.defined_names, 'keys') else wb.defined_names)


if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else None
    files = [SRC / target] if target else list(SRC.glob("*.xlsx"))
    for f in files:
        inspect(f)
